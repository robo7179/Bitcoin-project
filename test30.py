import torch
import torch.nn as nn
from torch.nn import Linear
from torch_geometric.nn.conv import TransformerConv
from torch_geometric.nn import Set2Set
from torch_geometric.nn import BatchNorm
from tqdm import tqdm
import torch.nn.functional as F
from torch_geometric.nn import NNConv
import torch.optim as optim

import os
import csv
import datetime
from torch_geometric.data import Data

from sklearn import preprocessing
from openpyxl import Workbook

wb = Workbook()
ws = wb.active 

ws['A1'] = "Date"
ws['B1'] = "Loss"
##ws.append(["Bob", 25])

# Save the workbook


def get_num(s):
    s1=""
    s2=""
    ind=-1
    for i in range(len(s)):
        if(s[i]=='E'):
            ind=i
            break
        else:
            s1=s1+s[i]
    if(ind==-1):
        return float(s)
    else:
        for i in range(ind+2,len(s)):
            s2=s2+s[i]
        return float(s1)*pow(10,float(s2))

class GVAE(torch.nn.Module):
        
	def __init__(self,input_dim,hidden_dim,edge_dim,latent_embedding_dim,hidden_dim2):

		super(GVAE, self).__init__()
		self.hidden_dim2 = hidden_dim2
		self.input_dim = input_dim
		self.hidden_dim = hidden_dim
		self.edge_dim = edge_dim
		self.latent_embedding_dim = latent_embedding_dim

		#MLP unit
		nn_edge=nn.Sequential(
		nn.Linear(edge_dim, hidden_dim),
		nn.ReLU(),
		nn.Linear(hidden_dim, input_dim * hidden_dim)
		)

		#encoder layers:

		self.conv1 = NNConv(input_dim, hidden_dim, nn_edge)
		self.bn1 = BatchNorm(hidden_dim)
		self.conv2 = NNConv(hidden_dim, hidden_dim, nn_edge)
		self.bn2 = BatchNorm(hidden_dim)
		self.conv3 = NNConv(hidden_dim, hidden_dim, nn_edge)
		self.bn3 = BatchNorm(hidden_dim)
		self.conv4 = NNConv(hidden_dim, hidden_dim, nn_edge)
		self.bn4 = BatchNorm(hidden_dim)

		self.mu = nn.Linear(self.hidden_dim, self.latent_embedding_dim)
		self.logvar = nn.Linear(self.hidden_dim, self.latent_embedding_dim)
		
		
		
		
		
		

		#decoder layers:

		self.decoder=nn.Sequential(
		nn.Linear(self.latent_embedding_dim*2,self.hidden_dim2),
		nn.ReLU(),
		nn.Linear(self.hidden_dim2,self.hidden_dim2),
		nn.ReLU(),
		nn.Linear(self.hidden_dim2,self.hidden_dim2),
		nn.ReLU(),
		nn.Linear(self.hidden_dim2,self.hidden_dim2),
		nn.ReLU(),
		nn.Linear(self.hidden_dim2,self.edge_dim)
		)

	def innerprod(self,z):
		return torch.sigmoid(torch.matmul(z,z.t()))

	def edge_feature(self, z, edge_index):

		source_embeddings = z[edge_index[0]]  # Shape: [num_edges, latent_embedding_dim]
		target_embeddings = z[edge_index[1]]  # Shape: [num_edges, latent_embedding_dim]

		# Concatenate the source and target embeddings
		edge_embeddings = torch.cat([source_embeddings, target_embeddings], dim=1)  # Shape: [num_edges, latent_embedding_dim * 2]

		# Pass the edge embeddings through the decoder network
		decoded_edge_features = self.decoder(edge_embeddings)  # Shape: [num_edges, num_edge_features]

		return decoded_edge_features

	def forward(self, x, edge_index, edge_attr):
		# Encoder
		x = F.relu(self.conv1(x, edge_index, edge_attr))
		x = self.bn1(x)
		x = F.relu(self.conv2(x, edge_index, edge_attr))
		x = self.bn2(x)
		#print("x ",torch.min(x)," ",torch.max(x))
		x = F.relu(self.conv3(x, edge_index, edge_attr))
		x = self.bn3(x)
		#print("x2 ",torch.min(x)," ",torch.max(x))
		x = F.relu(self.conv4(x, edge_index, edge_attr))
		x = self.bn4(x)
		#print("x3 ",torch.min(x)," ",torch.max(x))
		# Latent space
		mu = self.mu(x)
		logvar = self.logvar(x)
		logvar=torch.clamp(logvar,max=10)
		
		z = self.reparameterize(mu, logvar)

		# Decode edge features
		decoded_edge_features = self.edge_feature(z, edge_index)
		
		return decoded_edge_features

	def reparameterize(self, mu, logvar):
		std = torch.exp(0.5 * logvar)
		eps = torch.randn_like(std)
		return mu + eps * std

	def vgae_loss1(model,decoded_edge_features, edge_attr, mu, logvar):
		# Mean Squared Error Loss for edge feature reconstruction
		mse_loss = F.mse_loss(decoded_edge_features,edge_attr,reduction = "mean")
		# KL Divergence Loss to regularize the latent space
		#print("mu",mu)
		#print("logvar ",logvar)
		#print("logvar dim ",logvar.size())
		#print("mu dim ",mu.size())
		#print(torch.sum(logvar))
		#print(torch.sum(mu.pow(2)))
		#print(torch.sum(logvar.exp()))
		#print("hey: ",torch.max(logvar))
		
		#kl_loss = -0.5 * torch.sum(1 + logvar - mu.pow(2) - logvar.exp())
		kl_loss= -0.5 * torch.sum(1 + logvar - mu.pow(2) - logvar.exp())
		#print("mse_loss :",mse_loss)
		#print("kl_loss :",kl_loss)
		ans = mse_loss + kl_loss
		#print("loss in vgae_loss1 :",ans)
		return ans


	def train1(model,optimizer, num_epochs,x,edge_index,edge_attr):
		loss1=[]
		for epoch in range(num_epochs):
			print("epoch ",epoch)
			# Extract input data from the batch

			  # Zero the gradients
			optimizer.zero_grad()
			# Forward pass through the model
			decoded_edge_features = model(x, edge_index, edge_attr)

			# Compute the loss
			mu = model.mu(x)
			logvar = model.logvar(x)
			logvar=torch.clamp(logvar, max=10)
			#print("mu :",mu)
			#print("logvar :",logvar)
			loss = model.vgae_loss1(decoded_edge_features, edge_attr, mu, logvar)
			loss1.append(int(loss))
			
			print("loss ",loss)
			if(loss<=0.05):
				break
			loss.backward()  # Backward pass to compute gradients
			#for name,param in model.named_parameters():
			#	print("gradient :",name," ",param.grad.norm())
			optimizer.step()


		with open("output1.txt", "w") as file:
    
  
			for item in loss1:
				file.write(str(item) + "\n")



			
		

	def test1(model,x,edge_index,edge_attr):
		model.eval()
		with torch.no_grad():
			# Forward pass through the model
			decoded_edge_features = model(x, edge_index, edge_attr)
			mu = model.mu(x)
			logvar = model.logvar(x)
			logvar=torch.clamp(logvar,max=10)
			loss = model.vgae_loss1(decoded_edge_features, edge_attr, mu, logvar)
			
		return loss.item()

input_dim=1
hidden_dim=1
edge_dim=2
latent_embedding_dim=1
hidden_dim2=2

folder_path="/media/user/New/Bitcoin_Research/model_saved30"
file_path=os.path.join(folder_path,"model_path.pth")
os.makedirs(folder_path,exist_ok=True)
model = torch.load(file_path)





si="/media/user/New/Bitcoin_Research/inputs1/blockchair_bitcoin_inputs_"
so="/media/user/New/Bitcoin_Research/outputs1/blockchair_bitcoin_outputs_"
s_ext=".tsv"

start=datetime.date(2009,1,12)
end=datetime.date(2018,6,10)
res_date=start
b=False
slid_size=30
res_before=start
while(res_date<=end):
	sind=0
	res_date1=res_date
	print("on date: ",res_date)
	res_date += datetime.timedelta(days=15)
	dict_out={}
	dict_in={}
	node_ind=0
	dict_nodes={}
	dict_edges={}
	dict_nodes2={}
	dict_netdeg={}

	while(sind<slid_size and res_date1<=end):
		ss=res_date1.strftime("%Y%m%d")
		fsi=si+ss+s_ext
		
		#print("indexx ",sind)
		#print(fsi)
		fso=so+ss+s_ext
		#print(fso)
		res_date1 += datetime.timedelta(days=1)
		if(not os.path.exists(fsi)):#check if it works
			continue
		if(not os.path.exists(fso)):
			continue
		sind += 1
		
		csv_file_out=open(fso)
		csv_file_in=open(fsi)
		csv_reader_out=csv.reader(csv_file_out,delimiter="\t")
		csv_reader_in=csv.reader(csv_file_in,delimiter="\t")
		line_count=0
		#dict_out={}
		#dict_in={}

		for row in csv_reader_out:
			#print("row from output")
			#print(row)
			if(row[9]=='0'):
				thash=row[1]
				if(thash in dict_out):
					l=dict_out[thash]
					l.append([get_num(row[4])/(100000000),row[6]])
					#l.append([get_num(row[4]),row[6]])
					dict_out[thash]=l
				else:
					dict_out[thash]=[[get_num(row[4])/(100000000),row[6]]]
					#dict_out[thash]=[[get_num(row[4]),row[6]]]

		for row in csv_reader_in:
			if(row[4]=="value"):
				continue
			# print("row: ")
			#print(row)
			thash=row[12]
			if(thash in dict_in):
				l=dict_in[thash]
				l.append([(get_num(row[4])/100000000) , row[6] ])
				#l.append([get_num(row[4]),row[6]])
				dict_in[thash]=l
			else:
				dict_in[thash]=[[ (get_num(row[4])/100000000) , row[6] ]]
				#dict_in[thash]=[[ (get_num(row[4])) , row[6] ]]
			#node_ind=0
			#dict_nodes={}
			#dict_edges={}
			#dict_nodes2={}
		for key in dict_out:
			lo=dict_out[key]
			li=dict_in[key]
			lo.sort(reverse=True)
			li.sort(reverse=True)
			i=j=0
			while(len(lo)>0 and len(li)>0):
				lo.sort(reverse=True)
				li.sort(reverse=True)
				r=lo[i][1]
				s=li[j][1]
				if(s not in dict_nodes):
					dict_nodes[s]=node_ind
					dict_nodes2[node_ind]=s
					dict_netdeg[s]=0
					node_ind=node_ind+1
				if(r not in dict_nodes):
					dict_nodes[r]=node_ind
					dict_nodes2[node_ind]=r
					dict_netdeg[r]=0
					node_ind=node_ind+1

				if(lo[i][0]==li[j][0]):
					if((dict_nodes[s],dict_nodes[r]) not in dict_edges):
						dict_edges[(dict_nodes[s],dict_nodes[r])]=[li[j][0],1]
						dict_netdeg[s]=dict_netdeg[s]-1
						dict_netdeg[r]= dict_netdeg[r]+1
					else:
						l=dict_edges[(dict_nodes[s],dict_nodes[r])]
						l[0]+=li[j][0]
						l[1]=l[1]+1
						dict_edges[(dict_nodes[s],dict_nodes[r])]=l
					lo.pop(0)
					li.pop(0)
					#i=i+1
					#j=j+1
				elif(lo[i]<li[j]):

					if((dict_nodes[s],dict_nodes[r]) not in dict_edges):
						dict_edges[(dict_nodes[s],dict_nodes[r])]=[lo[i][0],1]
						dict_netdeg[s]=dict_netdeg[s]-1
						dict_netdeg[r]= dict_netdeg[r]+1
					else:
						l=dict_edges[(dict_nodes[s],dict_nodes[r])]
						l[0]+=lo[i][0]
						l[1]=l[1]+1
						dict_edges[(dict_nodes[s],dict_nodes[r])]=l
					li[j][0]-=lo[i][0]
					lo.pop(0)
					#i=i+1
				else:
					if((dict_nodes[s],dict_nodes[r]) not in dict_edges):
						dict_edges[(dict_nodes[s],dict_nodes[r])]=[li[j][0],1]
						dict_netdeg[s]=dict_netdeg[s]-1
						dict_netdeg[r]= dict_netdeg[r]+1
					else:
						l=dict_edges[(dict_nodes[s],dict_nodes[r])]
						l[0]+=li[j][0]
						l[1]=l[1]+1
						dict_edges[(dict_nodes[s],dict_nodes[r])]=l
					lo[i][0]-=li[j][0]
					li.pop(0)
					#j=j+1
	x=[]
	edge_index=[]
	edge_attr=[]
	#for (k,v) in dict_netdeg.items():
		#print(k,v)
	for i in range(node_ind):
		x.append([dict_netdeg[dict_nodes2[i]]])
	e1=[]
	e2=[]
	edge_features1=[]
	edge_features2=[]
	maxi=0
	for key,val in dict_edges.items():
		#print(key,val)
		e1.append(key[0])
		e2.append(key[1])
		edge_attr.append([val[0],val[1]])
		maxi=max(maxi,val[1])
		edge_features1.append([val[0]])
		edge_features2.append([val[1]])
	#print("maximun no of transactions: ",maxi)
	min_max_scaler = preprocessing.MinMaxScaler(feature_range =(0, 1))
	if(len(edge_features1)==0):
		continue
	min_max_scaled_edge_features1=min_max_scaler.fit_transform(edge_features1)
	min_max_scaled_edge_features2=min_max_scaler.fit_transform(edge_features2)
	#print("maaximum of edge_features1 scaled: ",max(min_max_scaled_edge_features1)," ",min(min_max_scaled_edge_features1))
	#print("maaximum of edge_features2 scaled: ",max(min_max_scaled_edge_features2)," ",min(min_max_scaled_edge_features2))
	for i in range(len(edge_attr)):
		edge_attr[i][0]=min_max_scaled_edge_features1[i][0]
		edge_attr[i][1]=min_max_scaled_edge_features2[i][0]
	edge_index.append(e1)
	edge_index.append(e2)
	
	x=torch.tensor(x,dtype=torch.float)
	edge_index=torch.tensor(edge_index,dtype=torch.long)
	edge_attr=torch.tensor(edge_attr,dtype=torch.float)
	data = Data(x=x, edge_index=edge_index, edge_attr=edge_attr)
	#print("going for train")
	#def test1(model,x,edge_index,edge_attr):
	loss=model.test1(data.x,data.edge_index,data.edge_attr)
	print("loss",loss)
	with open("output1.txt", "w") as file:
		file.write(str(res_date)+" : "+str(loss) + "\n")
	ws.append([str(res_date),str(loss)])
	wb.save("test_profile30.xlsx")
	#print("date :",res_date)
	#print(x,edge_index,edge_attr)
	
	output = model(x, edge_index, edge_attr)
	#print(output)

	

print("Model succesfully saved")
wb.save("test_profile30.xlsx")
	
	
	

